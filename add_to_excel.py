# /Users/shubham.thorat/Downloads/Load_Test.xlsx
from openpyxl import load_workbook
import json
# No	connections	threads	duration(ms)	req/sec	bytes_trasfer/sec	50th percentile	90th percentile	99th percentile																	
def convertJSONToArray(index,item):
    try:
      rows = [
          index,
          item['maxRequests'],
          item['concurrency'],
          item['concurrentClients'],
          item['completedRequests'],
          item['errors'],
          item['totalTime'],
          item['meanLatency'],
          item['effectiveRPS'],
          item['effectiveRPSPerClient'],
          item['50th_percentile'],
          item['90th_percentile'],
          item['99th_percentile'],
          item['100th_percentile']
      ]

      return rows
    except Exception as e:
      print("Error while parsing json data : ",e)

def addJSONToExcel(inputJSONFile, outputXlsxFile):
    try:
        jsonFile = open(inputJSONFile)
        data = json.load(jsonFile)
        wb = load_workbook(outputXlsxFile)
        sheet_name = 'WS_2'  # Change this to the desired sheet name
        sheet = wb[sheet_name]
        index = sheet.max_row + 1
        print('Starting row : ',index)
        sheet.append([])
        index += 1
        for item in data:
            # print("item = ",item)
            if not item:
                pass
            try:
                row = convertJSONToArray(index,item)
                sheet.append(row)
                index += 1
                print(f'Row added success : ',index)
            except Exception as e:
                print(f"Error while appending row to excel row={index} {e}")

        wb.save(outputXlsxFile)
    except Exception as e:
        print(f"An error occurred addJSONToExcel: {e}")


def main():
    inputJSONFile = 'ws_output.json'
    path = 'Load_Test_3.xlsx'
    addJSONToExcel(inputJSONFile=inputJSONFile,outputXlsxFile=path)

main()